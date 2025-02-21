import io
from datetime import datetime
from typing import Any, Dict, List, Optional, Tuple, Union

import pandas as pd
from django.http import HttpResponse
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from utils.log.logger import logger


def style_excel_cell(
    cell: Any,
    font_size: int = 11,
    bold: bool = False,
    align: str = "center",
    bg_color: Optional[str] = None,
    font_color: Optional[str] = None,
    border: bool = True,
) -> None:
    """
    设置Excel单元格样式
    :param cell: 单元格对象
    :param font_size: 字体大小
    :param bold: 是否加粗
    :param align: 对齐方式
    :param bg_color: 背景颜色
    :param font_color: 字体颜色
    :param border: 是否添加边框
    """
    # 设置字体
    cell.font = Font(
        name="微软雅黑",
        size=font_size,
        bold=bold,
        color=font_color,
    )

    # 设置对齐
    cell.alignment = Alignment(
        horizontal=align,
        vertical="center",
        wrap_text=True,
    )

    # 设置背景色
    if bg_color:
        cell.fill = PatternFill(
            start_color=bg_color,
            end_color=bg_color,
            fill_type="solid",
        )

    # 设置边框
    if border:
        border_style = Side(style="thin", color="000000")
        cell.border = Border(
            left=border_style,
            right=border_style,
            top=border_style,
            bottom=border_style,
        )


def auto_adjust_columns(worksheet: Worksheet, min_width: int = 10, max_width: int = 50) -> None:
    """
    自动调整列宽
    :param worksheet: 工作表对象
    :param min_width: 最小列宽
    :param max_width: 最大列宽
    """
    for column in worksheet.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)

        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass

        adjusted_width = (max_length + 2) * 1.2
        worksheet.column_dimensions[column_letter].width = max(
            min(adjusted_width, max_width),
            min_width,
        )


def create_excel_workbook(
    data: List[Dict[str, Any]],
    headers: Optional[Dict[str, str]] = None,
    sheet_name: str = "Sheet1",
) -> Workbook:
    """
    创建Excel工作簿
    :param data: 数据列表
    :param headers: 表头映射
    :param sheet_name: 工作表名称
    :return: 工作簿对象
    """
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = sheet_name

    # 如果没有提供表头映射，使用数据的键作为表头
    if not headers and data:
        headers = {key: key for key in data[0].keys()}

    # 写入表头
    if headers:
        for col, header in enumerate(headers.values(), 1):
            cell = worksheet.cell(row=1, column=col, value=header)
            style_excel_cell(cell, bold=True, bg_color="CCCCCC")

        # 写入数据
        for row, item in enumerate(data, 2):
            for col, key in enumerate(headers.keys(), 1):
                cell = worksheet.cell(row=row, column=col, value=item.get(key))
                style_excel_cell(cell)

    # 调整列宽
    auto_adjust_columns(worksheet)

    return workbook


def pandas_read_excel(
    file_path: str,
    sheet_name: Optional[Union[str, int]] = 0,
    header: Optional[Union[int, List[int]]] = 0,
    names: Optional[List[str]] = None,
    skiprows: Optional[Union[int, List[int]]] = None,
    na_values: Optional[List[str]] = None,
) -> pd.DataFrame:
    """
    使用pandas读取Excel文件
    :param file_path: 文件路径
    :param sheet_name: 工作表名称或索引
    :param header: 表头行号
    :param names: 列名列表
    :param skiprows: 跳过的行号
    :param na_values: 空值替换列表
    :return: DataFrame对象
    """
    try:
        df = pd.read_excel(
            file_path,
            sheet_name=sheet_name,
            header=header,
            names=names,
            skiprows=skiprows,
            na_values=na_values or ["", "NA", "N/A", "null", "NULL", "none", "None"],
        )
        return df
    except Exception as e:
        logger.error(f"读取Excel文件失败: {str(e)}")
        raise


def pandas_write_excel(
    df: pd.DataFrame,
    file_path: str,
    sheet_name: str = "Sheet1",
    index: bool = False,
    header: bool = True,
    encoding: str = "utf-8",
) -> None:
    """
    使用pandas写入Excel文件
    :param df: DataFrame对象
    :param file_path: 文件路径
    :param sheet_name: 工作表名称
    :param index: 是否包含索引
    :param header: 是否包含表头
    :param encoding: 编码格式
    """
    try:
        df.to_excel(
            file_path,
            sheet_name=sheet_name,
            index=index,
            header=header,
            encoding=encoding,
        )
    except Exception as e:
        logger.error(f"写入Excel文件失败: {str(e)}")
        raise


def pandas_download_excel(
    data: Union[pd.DataFrame, List[Dict[str, Any]]],
    filename: Optional[str] = None,
    sheet_name: str = "Sheet1",
    index: bool = False,
) -> HttpResponse:
    """
    下载Excel文件
    :param data: DataFrame对象或数据列表
    :param filename: 文件名
    :param sheet_name: 工作表名称
    :param index: 是否包含索引
    :return: HTTP响应对象
    """
    try:
        # 如果是数据列表，转换为DataFrame
        if isinstance(data, list):
            df = pd.DataFrame(data)
        else:
            df = data

        # 创建一个字节流
        output = io.BytesIO()

        # 写入Excel
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            df.to_excel(writer, sheet_name=sheet_name, index=index)

            # 获取工作表
            worksheet = writer.sheets[sheet_name]

            # 设置样式
            for row in worksheet.iter_rows(min_row=1, max_row=1):
                for cell in row:
                    style_excel_cell(cell, bold=True, bg_color="CCCCCC")

            for row in worksheet.iter_rows(min_row=2):
                for cell in row:
                    style_excel_cell(cell)

            # 调整列宽
            auto_adjust_columns(worksheet)

        # 设置响应头
        if not filename:
            filename = f"export_{datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"

        response = HttpResponse(
            output.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response

    except Exception as e:
        logger.error(f"导出Excel文件失败: {str(e)}")
        raise


def validate_excel_template(
    file_path: str,
    required_headers: List[str],
    sheet_name: Optional[Union[str, int]] = 0,
) -> Tuple[bool, Optional[str]]:
    """
    验证Excel模板
    :param file_path: 文件路径
    :param required_headers: 必需的表头列表
    :param sheet_name: 工作表名称或索引
    :return: (是否有效, 错误信息)
    """
    try:
        # 读取Excel文件
        df = pd.read_excel(file_path, sheet_name=sheet_name, nrows=0)
        headers = df.columns.tolist()

        # 检查必需的表头
        missing_headers = [header for header in required_headers if header not in headers]
        if missing_headers:
            return False, f"缺少必需的表头: {', '.join(missing_headers)}"

        return True, None

    except Exception as e:
        logger.error(f"验证Excel模板失败: {str(e)}")
        return False, f"验证Excel模板失败: {str(e)}"


def process_excel_data(
    df: pd.DataFrame,
    column_mapping: Optional[Dict[str, str]] = None,
    converters: Optional[Dict[str, callable]] = None,
    validators: Optional[Dict[str, callable]] = None,
) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]]]:
    """
    处理Excel数据
    :param df: DataFrame对象
    :param column_mapping: 列名映射
    :param converters: 数据转换器
    :param validators: 数据验证器
    :return: (有效数据列表, 无效数据列表)
    """
    valid_data = []
    invalid_data = []

    try:
        # 重命名列
        if column_mapping:
            df = df.rename(columns=column_mapping)

        # 处理每一行数据
        for index, row in df.iterrows():
            row_data = row.to_dict()
            row_errors = []

            # 数据转换
            if converters:
                for field, converter in converters.items():
                    try:
                        row_data[field] = converter(row_data.get(field))
                    except Exception as e:
                        row_errors.append(f"{field}: {str(e)}")

            # 数据验证
            if validators:
                for field, validator in validators.items():
                    try:
                        if not validator(row_data.get(field)):
                            row_errors.append(f"{field}: 验证失败")
                    except Exception as e:
                        row_errors.append(f"{field}: {str(e)}")

            # 根据验证结果分类
            if row_errors:
                row_data["errors"] = row_errors
                invalid_data.append(row_data)
            else:
                valid_data.append(row_data)

        return valid_data, invalid_data

    except Exception as e:
        logger.error(f"处理Excel数据失败: {str(e)}")
        raise


"""
使用示例：

# 导出Excel
data = [
    {"name": "张三", "age": 20, "email": "zhangsan@example.com"},
    {"name": "李四", "age": 25, "email": "lisi@example.com"}
]
headers = {
    "name": "姓名",
    "age": "年龄",
    "email": "邮箱"
}
workbook = create_excel_workbook(data, headers)
workbook.save("example.xlsx")

# 读取Excel
df = pandas_read_excel("example.xlsx")
print(df)

# 下载Excel
response = pandas_download_excel(data, "example.xlsx")

# 验证Excel模板
valid, error = validate_excel_template(
    "template.xlsx",
    ["姓名", "年龄", "邮箱"]
)

# 处理Excel数据
column_mapping = {
    "姓名": "name",
    "年龄": "age",
    "邮箱": "email"
}
converters = {
    "age": int
}
validators = {
    "email": lambda x: "@" in str(x)
}
valid_data, invalid_data = process_excel_data(
    df,
    column_mapping,
    converters,
    validators
)
"""
